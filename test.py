import os
import re
import pandas as pd
from openpyxl import load_workbook
import shutil

def process_folder(new_data_folder_path,data_folder_path, output_df, audio_type, data_wb,words_df,sens_df):
    audio_prefix = "audio_w" if audio_type == "AUDIO WORD" else "audio_s"

    file_count = len([f for f in os.listdir(data_folder_path)])

    for filename in os.listdir(new_data_folder_path):
        if filename.endswith(".mp3") or filename.endswith(".m4a"):
            file_number = filename.split(".")[0]
            file_number = file_number.split("-")[0]

            file_number = file_number.split(" ")[0]
            new_file_number = int(file_number) + file_count + 30
            if filename.endswith(".mp3"):
                new_filename = f"{audio_prefix}{new_file_number}.mp3"
            else:
                new_filename = f"{audio_prefix}{new_file_number}.m4a"

            # Read data from words.xlsx and sens.xlsx

            # Get data from respective rows
            if audio_prefix == "audio_s":
                row_2 = sens_df.iloc[int(file_number) - 1, 1]
                row_3 = sens_df.iloc[int(file_number) - 1, 0]
            else:
                row_2 = words_df.iloc[int(file_number) - 1, 1]
                row_3 = words_df.iloc[int(file_number) - 1, 0]

            # Remove "12.", "13.",... at the beginning of row_3
            row_3 = re.sub(r'^\d+\.', '', row_3.strip())
            row_2 = re.sub(r'^\d+\.', '', row_2.strip())
            # Append data to the output DataFrame
            output_df = pd.concat([output_df, pd.DataFrame({"File": [new_filename], "Mnong Transcript": [row_2], "Transcript": [row_3]})])
            new_file_path = os.path.join(data_folder_path, new_filename)
            file_path = os.path.join(new_data_folder_path, filename)

            # Move the file to the 'New Data' folder
            shutil.move(file_path, new_file_path)
            print("moved",file_path,"to",new_file_path)

            # Check if sheet 2 exists, otherwise create it
            if "Sheet1" not in data_wb.sheetnames:
                sheet1 = data_wb.create_sheet(title="Sheet1")
                sheet1.append(["File", "Mnong Transcript", "Transcript"])
            else:
                sheet1 = data_wb["Sheet1"]

            # Add data to sheet 2 of the Excel file
            sheet1.append([new_filename, row_2, row_3])
            data_wb.save("data.xlsx")
    return output_df

def main():
    new_data_folder = "New Data"
    data_folder = "Data"
    output_df = pd.DataFrame(columns=["File", "Mnong Transcript", "Transcript"])
    words_df = pd.read_excel(os.path.join('Data Detail', 'Viet_MNong_TU VUNG.xlsx'), header=None)
    sens_df = pd.read_excel(os.path.join('Data Detail', 'Dịch Tiếng Mnông T2.2024 Tuyn (1).xlsx'), header=None)
    # Load existing Excel workbook
    data_wb_path = "data.xlsx"
    if os.path.exists(data_wb_path):
        data_wb = load_workbook(data_wb_path)
    else:
        data_wb = load_workbook()

    for subfolder in os.listdir(new_data_folder):
        subfolder_path = os.path.join(new_data_folder, subfolder)
        if os.path.isdir(subfolder_path):
            output_df = process_folder(subfolder_path,data_folder, output_df, subfolder, data_wb,words_df,sens_df)

    # Save the result to data.xlsx
    data_wb.save(data_wb_path)

if __name__ == "__main__":
    main()
